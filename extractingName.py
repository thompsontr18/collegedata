from bs4 import BeautifulSoup
import requests
import pandas as pd
from googlesearch import search
import re
from openpyxl import load_workbook
import time



def googlesearch(query):
    time.sleep(0.01)
    for j in search(query, tld="com", stop=1, pause=2):
        return j



df = pd.read_excel('data.xlsx', usecols='A,B,C')
df_print = pd.DataFrame(columns=['Email IDs','Other Details'])
book = load_workbook('dataWithEmail.xlsx')
sheet=book.worksheets[0]

major = ""
collegeandmajor = {}


#for i in range(len(df)):
for i in range(18000, len(df)):
    srch = str(df.iloc[i, 0]) + " " + str(df.iloc[i, 1])
    if df.iloc[i, 0] == "Wiregrass Georgia Technical College":
        df_print.loc[len(df_print.index)] = ["N/A",""]
        continue
    # if df.iloc[i, 1] == "ACCT" and df.iloc[i, 0] == "Wiregrass Georgia Technical College":
    #     df_print.loc[len(df_print.index)] = ["N/A",""]
    #     continue
    srch += " program contact"
    if srch in collegeandmajor:
        df_print.loc[len(df_print.index)] = collegeandmajor[srch]
        continue
    print(df.iloc[i, 0])
    #print(srch)
    link = googlesearch(srch)
    if major != df.iloc[i, 1]:
        print('-----------------------------------')
        print(df.iloc[i, 1])
        major = df.iloc[i, 1]
        # collegeandmajor[(df.iloc[i, 0], df.iloc[i, 1])] = 1

    time.sleep(0.01)
    html = requests.get(link, verify=False)
    soup = BeautifulSoup(html.text, "html.parser")
    text = soup.text

    emailre = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'
    emailmatches = []
    emailmatches = re.findall(emailre, text)


    strings = ""
    for i in range(len(text)):
        # strings += text[i]
        if "@" in text[i]:
            # strings = strings[-150:]
            strings = text[i-150:i+50]
            #print(strings)


    if len(emailmatches) == 0:
        #print("N/A")
        df_print.loc[len(df_print.index)] = ["N/A",strings]
        collegeandmajor[srch] = ["N/A",strings]
        continue
    printedtext = ""
    for i in range(0,len(emailmatches)):
        if i != len(emailmatches)-1:
            printedtext += emailmatches[i] + ", "
        else:
            printedtext += emailmatches[i]
    #print(printedtext)
    df_print.loc[len(df_print.index)] = [printedtext,strings]
    
    collegeandmajor[srch] = [printedtext,strings]

df_print = df_print.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
with pd.ExcelWriter('dataWithEmail.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df_print.to_excel(writer,sheet_name="Sheet1",header=False, index=False, startrow=sheet.max_row)




