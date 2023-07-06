import pandas as pd
import re
from openpyxl import load_workbook



df = pd.read_excel('data.xlsx', usecols='H')
df_print = pd.DataFrame(columns=['Phone Number'])
book = load_workbook('datawithPhone.xlsx')
sheet=book.worksheets[0]


phonenumber = r'\(?\d{3}\)?.?[-\s]?\d{3}.?[-\s]?\d{4}'
phonenumbermatches = []


for i in range(5000, len(df)):
    cell = str(df.iloc[i, 0])
    phonenumbermatches = re.findall(phonenumber, cell)
    #print(phonenumbermatches)
    if len(phonenumbermatches) == 0:
        print("N/A")
        df_print.loc[len(df_print.index)] = ["N/A"]
    else:
        stri = ""
        for i in range(0,len(phonenumbermatches)):
            if i != len(phonenumbermatches)-1:
                stri += phonenumbermatches[i] + ", "
            else:
                stri += phonenumbermatches[i]
        print(stri)
        df_print.loc[len(df_print.index)] = [stri]

df_print = df_print.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
with pd.ExcelWriter('datawithPhone.xlsx',mode='a', if_sheet_exists='overlay') as writer:  
    df_print.to_excel(writer,sheet_name="Sheet1",header=True, index=False, startrow=sheet.max_row)


